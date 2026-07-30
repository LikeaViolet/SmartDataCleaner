from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.worksheet.worksheet import Worksheet


STATUS_SUFFIX = " Status"

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="F04F64",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

MISSING_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF4D6",
)

INVALID_FILL = PatternFill(
    fill_type="solid",
    fgColor="FDE2E2",
)

DUPLICATE_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF1F3",
)


def dataframe_to_csv_bytes(
    df: pd.DataFrame,
) -> bytes:
    """Convert a dataframe into downloadable CSV bytes."""

    return df.to_csv(
        index=False
    ).encode("utf-8")


def _business_columns(
    dataframe: pd.DataFrame,
) -> list[str]:
    return [
        column
        for column in dataframe.columns
        if not str(column).endswith(
            STATUS_SUFFIX
        )
    ]


def build_cleaned_data_export(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """
    Return client-ready data without internal status columns.
    """

    columns = _business_columns(dataframe)

    return dataframe.loc[:, columns].copy()


def build_quality_issues_export(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """
    Build a cell-level list of missing and invalid values.
    """

    issue_columns = [
        "Data row",
        "Column",
        "Value",
        "Status",
    ]

    business_columns = _business_columns(
        dataframe
    )

    status_columns = [
        column
        for column in dataframe.columns
        if str(column).endswith(
            STATUS_SUFFIX
        )
    ]

    issues: list[dict[str, object]] = []
    recorded_issues: set[
        tuple[int, str]
    ] = set()

    for row_number, (_, row) in enumerate(
        dataframe.iterrows(),
        start=2,
    ):
        for status_column in status_columns:
            status_value = row[status_column]

            if pd.isna(status_value):
                continue

            normalized_status = (
                str(status_value)
                .strip()
                .lower()
            )

            if normalized_status in {
                "valid",
                "standardized",
            }:
                continue

            source_column = str(
                status_column
            )[
                :-len(STATUS_SUFFIX)
            ]

            source_value = (
                row[source_column]
                if source_column in row.index
                else pd.NA
            )

            issues.append(
                {
                    "Data row": row_number,
                    "Column": source_column,
                    "Value": source_value,
                    "Status": str(
                        status_value
                    ),
                }
            )

            recorded_issues.add(
                (
                    row_number,
                    source_column,
                )
            )

        for column in business_columns:
            value = row[column]

            is_missing = (
                pd.isna(value)
                or (
                    isinstance(value, str)
                    and not value.strip()
                )
            )

            issue_key = (
                row_number,
                str(column),
            )

            if (
                is_missing
                and issue_key
                not in recorded_issues
            ):
                issues.append(
                    {
                        "Data row": row_number,
                        "Column": str(column),
                        "Value": value,
                        "Status": "Missing",
                    }
                )

                recorded_issues.add(
                    issue_key
                )

    return pd.DataFrame(
        issues,
        columns=issue_columns,
    )


def _style_worksheet(
    worksheet: Worksheet,
) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

    worksheet.row_dimensions[1].height = 24

    for column_cells in worksheet.columns:
        column_letter = (
            column_cells[0].column_letter
        )

        maximum_length = max(
            (
                len(str(cell.value))
                if cell.value is not None
                else 0
            )
            for cell in column_cells
        )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(maximum_length + 2, 12),
            40,
        )


def _style_quality_issues(
    worksheet: Worksheet,
) -> None:
    status_column = None

    for cell in worksheet[1]:
        if cell.value == "Status":
            status_column = cell.column
            break

    if status_column is None:
        return

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        status_cell = worksheet.cell(
            row=row_number,
            column=status_column,
        )

        status = str(
            status_cell.value or ""
        ).strip().lower()

        if status == "missing":
            status_cell.fill = MISSING_FILL

        elif status == "invalid":
            status_cell.fill = INVALID_FILL


def _style_removed_duplicates(
    worksheet: Worksheet,
) -> None:
    for row in worksheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.fill = DUPLICATE_FILL


def dataframe_to_excel_bytes(
    df: pd.DataFrame,
    removed_duplicates: (
        pd.DataFrame | None
    ) = None,
) -> bytes:
    """
    Create a professional Excel workbook containing cleaned
    data, quality issues, and removed duplicate records.
    """

    cleaned_export = (
        build_cleaned_data_export(df)
    )

    quality_issues = (
        build_quality_issues_export(df)
    )

    if removed_duplicates is None:
        removed_duplicates = pd.DataFrame(
            columns=cleaned_export.columns
        )
    else:
        removed_duplicates = (
            removed_duplicates.copy()
        )

    buffer = BytesIO()

    with pd.ExcelWriter(
        buffer,
        engine="openpyxl",
    ) as writer:
        cleaned_export.to_excel(
            writer,
            index=False,
            sheet_name="Cleaned Data",
        )

        quality_issues.to_excel(
            writer,
            index=False,
            sheet_name="Quality Issues",
        )

        removed_duplicates.to_excel(
            writer,
            index=False,
            sheet_name="Removed Duplicates",
        )

        workbook = writer.book

        cleaned_sheet = workbook[
            "Cleaned Data"
        ]

        issues_sheet = workbook[
            "Quality Issues"
        ]

        duplicates_sheet = workbook[
            "Removed Duplicates"
        ]

        _style_worksheet(cleaned_sheet)
        _style_worksheet(issues_sheet)
        _style_worksheet(duplicates_sheet)

        _style_quality_issues(
            issues_sheet
        )

        _style_removed_duplicates(
            duplicates_sheet
        )

    buffer.seek(0)

    return buffer.getvalue()